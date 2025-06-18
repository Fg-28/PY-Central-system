import os
import json
import time
import random
import tkinter as tk
from tkinter import messagebox, filedialog
import base64
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from shutil import copyfile
from collections import defaultdict

# === FILE PICKER ===
file_path = filedialog.askopenfilename(
    title="Select Excel File",
    filetypes=[("Excel files", "*.xlsx *.xlsm")]
)

if not file_path:
    messagebox.showerror("Error", "No file selected.")
    exit()

folder = os.path.dirname(file_path)
base = os.path.splitext(os.path.basename(file_path))[0]
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
new_filename = f"{base}_{timestamp}.xlsx"
new_path = os.path.join(folder, new_filename)
copyfile(file_path, new_path)

# === STEP 0: Fake Initial Loading ===
def show_fake_window(title, message, duration):
    window = tk.Tk()
    window.title(title)
    window.geometry("350x100+500+300")
    window.resizable(False, False)
    label = tk.Label(window, text=message, font=("Segoe UI", 12), padx=20, pady=20)
    label.pack()
    window.attributes("-topmost", True)
    window.overrideredirect(True)
    window.update()
    window.after(duration * 1000, window.destroy)
    window.mainloop()

show_fake_window("Processing", "🔄 File processing... Please wait...", 15)
show_fake_window("Encrypting", "🔐 Encrypting execution code...", 7)

# === STEP 1: Excel GST Reco ===
wb = load_workbook(new_path)
ws = wb.active
last_row = ws.max_row
last_col = ws.max_column
invoice_col = 2

invoice_numbers = [ws.cell(row=i, column=invoice_col).value for i in range(2, last_row + 1)]
duplicates = set(x for x in invoice_numbers if invoice_numbers.count(x) > 1)

ws_match = wb.create_sheet("Match")
for col in range(1, last_col + 1):
    ws_match.cell(row=1, column=col).value = ws.cell(row=1, column=col).value

match_row = 2
for i in range(2, last_row + 1):
    if ws.cell(row=i, column=invoice_col).value in duplicates:
        for col in range(1, last_col + 1):
            ws_match.cell(row=match_row, column=col).value = ws.cell(row=i, column=col).value
        match_row += 1

for i in range(last_row, 1, -1):
    if ws.cell(row=i, column=invoice_col).value in duplicates:
        ws.delete_rows(i)

taxable_col = 5
pivot_data = defaultdict(float)
for i in range(2, ws_match.max_row + 1):
    inv = ws_match.cell(row=i, column=invoice_col).value
    tax = ws_match.cell(row=i, column=taxable_col).value
    if isinstance(tax, (int, float)):
        pivot_data[inv] += tax

ws_pivot = wb.create_sheet("pivot1")
ws_pivot["E1"] = "INVOICE NUMBER"
ws_pivot["F1"] = "Sum of TAXABLE"
pivot_row = 2
for inv, total in pivot_data.items():
    if abs(total) >= 10:
        ws_pivot[f"E{pivot_row}"] = inv
        ws_pivot[f"F{pivot_row}"] = total
        pivot_row += 1

for i in range(2, pivot_row):
    ws_match.cell(row=i - 1, column=15).value = ws_pivot.cell(row=i, column=5).value

green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
highlight_vals = {ws_match.cell(row=r, column=15).value for r in range(1, ws_match.max_row + 1)}

for i in range(2, ws_match.max_row + 1):
    val = ws_match.cell(row=i, column=2).value
    if val in highlight_vals:
        ws_match.cell(row=i, column=2).fill = green_fill

ws_amt = wb.create_sheet("Amt Diff")
headers = ["DATE", "INVOICE NUMBER", "NAME", "GST NUMBER", "TAXABLE", "ABS", "IGST", "CGST", "SGST", "TOTAL", "STATUS"]
for col, header in enumerate(headers, 1):
    ws_amt.cell(row=1, column=col).value = header

row_amt = 2
for i in range(2, ws_match.max_row + 1):
    if ws_match.cell(row=i, column=2).fill == green_fill:
        for col in range(1, 12):
            ws_amt.cell(row=row_amt, column=col).value = ws_match.cell(row=i, column=col).value
        row_amt += 1

for i in range(ws_match.max_row, 1, -1):
    if ws_match.cell(row=i, column=2).fill == green_fill:
        ws_match.delete_rows(i)

match_rows = sorted(ws_match.iter_rows(min_row=2, max_col=11, values_only=True), key=lambda x: str(x[1]))
amt_rows = sorted(ws_amt.iter_rows(min_row=2, max_col=11, values_only=True), key=lambda x: str(x[1]))

for i, row in enumerate(match_rows, start=2):
    for j, val in enumerate(row, start=1):
        ws_match.cell(row=i, column=j).value = val

for i, row in enumerate(amt_rows, start=2):
    for j, val in enumerate(row, start=1):
        ws_amt.cell(row=i, column=j).value = val

ws_amt.auto_filter.ref = f"A1:K{ws_amt.max_row}"

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
for sheet_name in ["Amt Diff", "Match", "Sheet1"]:
    if sheet_name in wb.sheetnames:
        ws_target = wb[sheet_name]
        for row in range(2, ws_target.max_row + 1):
            cell_val = ws_target.cell(row=row, column=11).value
            if isinstance(cell_val, str) and ("GST R2B" in cell_val or "GSTR1" in cell_val or "GST R2A" in cell_val):
                for col in range(1, 12):
                    ws_target.cell(row=row, column=col).fill = yellow_fill

wb.save(new_path)

# === FINAL COMPLETION DIALOGS ===
root = tk.Tk()
root.withdraw()
messagebox.showinfo("Verification", "✅ Verification Complete")
messagebox.showinfo("Process", "✅ File Process Complete")
